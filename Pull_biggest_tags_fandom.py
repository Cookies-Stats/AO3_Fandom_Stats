import time
import requests
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import Workbook

# Call user header from .txt. AO3 asks that bots/scrapers have user agent headings. There are ways to search online for your user agent.
with open("user_agent.txt", "r") as file:
    user_agent = file.read().strip()

headers = {
    "User-Agent": user_agent
}

# Function to login to AO3
def login(username, password):
    login_url = "https://archiveofourown.org/users/login"
    session = requests.Session()
    response = session.get(login_url)
    soup = BeautifulSoup(response.text, "html.parser")

    authenticity_token = soup.find("input", {"name": "authenticity_token"})["value"]

    login_data = {
        "user[login]": username,
        "user[password]": password,
        "authenticity_token": authenticity_token,
        "commit": "Log in",
    }
    
    session.post(login_url, data=login_data)
    return session

# Define the function to get the tag data
def get_tag_data(session, url):
	response = session.get(url)
	soup = BeautifulSoup(response.text, "html.parser")
	main_div = soup.find("div", {"class": "child listbox group"})
	
	data = []

	if main_div:
		headings = main_div.find_all("h4", class_="heading")
		tag_lists = main_div.find_all("ul", class_="tags commas index group")
		
		for heading, tag_list in zip(headings, tag_lists):
			category = heading.get_text(strip=True)
			tags = [tag.get_text(strip=True) for tag in tag_list.find_all("a", class_="tag")]
			data.append((category, tags))
			
	return data
	
# Main script
def main():
	# Read credentials from the configuration file
	with open("login.txt", "r") as config_file:
		username = config_file.readline().strip()
		password = config_file.readline().strip()

	# Log in to AO3
	session = login(username, password)

	# Request tag from user
	input_tag = input("Please enter the specific tag name:")
	tag_url = f"https://archiveofourown.org/tags/{input_tag}"
	
	tag_data = get_tag_data(session, tag_url)
	
	workbook = Workbook()
	worksheet = workbook.active
	worksheet.title = "Results"
	
	worksheet.cell(row=1, column=1, value = "Category")
	worksheet.cell(row=1, column=2, value= "Tags")
	
	for index, (category, tags) in enumerate(tag_data, start=2):
		worksheet.cell(row=index, column=1, value=category)
		worksheet.cell(row=index, column=2, value=", ".join(tags))

    	
		time.sleep(10)  # Rate limiting per AO3's requests (can be as low as 5, 10 used for extra courtesy)
		
		
	# Save the workbook to an Excel file
	workbook.save("tag_results.xlsx")
	print("Data successfully exported")

# Run the script
if __name__ == "__main__":
    main()