import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
import re
import time

# Call user header from .txt. AO3 asks that bots/scrapers have user agent headings. There are ways to search online for your user agent.
with open("user_agent.txt", "r") as file:
    user_agent = file.read().strip()

headers = {
    "User-Agent": user_agent
}

# Function to login to AO3
def login(username, password):
    login_url = "https://archiveofourown.org/users/login"
    session = requests.Session()
    response = session.get(login_url)
    soup = BeautifulSoup(response.text, "html.parser")

    authenticity_token = soup.find("input", {"name": "authenticity_token"})["value"]

    login_data = {
        "user[login]": username,
        "user[password]": password,
        "authenticity_token": authenticity_token,
        "commit": "Log in",
    }
    
    session.post(login_url, data=login_data)
    return session

# Function to extract the number of found works
def get_work_count(session, url):
    response = session.get(url)
    soup = BeautifulSoup(response.text, "html.parser")
    
    main_div = soup.find("div", {"id": "main"})
    
    if main_div:
        header = main_div.find("h3", class_="heading")
        if header:
            found_text = header.get_text(strip=True)
            return found_text
    return None


# Main script
def main():
    # Read credentials from the configuration file
    with open("login.txt", "r") as config_file:
        username = config_file.readline().strip()
        password = config_file.readline().strip()

    # Log in to AO3
    session = login(username, password)

    # Read URLs from the text file
    with open("urls.txt", "r", encoding='utf-8') as file:
        lines = file.readlines()
		
    data = []
	
    for line in lines:
        #Split the line into three columns: Pairing, Category, and URL
        pairing, category, url = line.strip().split("\t")
	
	#Add rate limiting here
	time.sleep(5) # Rate limiting per AO3's requests (can be as low as 5 and can be higher)
		
        work_count = get_work_count(session, url)
        data.append((pairing, category, work_count))
        print(f"Pairing: {pairing}, Category: {category}, Works: {work_count}")

    # Create a new Excel workbook
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Results"

    # Write header to the Excel worksheet
    worksheet.cell(row=1, column=1, value="Pairing")
    worksheet.cell(row=1, column=2, value="Category")
    worksheet.cell(row=1, column=3, value="Work Count")
	

    # Loop through URLs and write the results to the Excel worksheet
    for index, (pairing, category, work_count) in enumerate(data, start=2):
        worksheet.cell(row=index, column=1, value=pairing)
        worksheet.cell(row=index, column=2, value=category)
        worksheet.cell(row=index, column=3, value=work_count)
		
		# Print the progress message
        print(f"URL {index - 1} successfully scraped")

    # Save the workbook to an Excel file
    workbook.save("results.xlsx")

# Run the script
if __name__ == "__main__":
    main()
