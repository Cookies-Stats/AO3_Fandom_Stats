import openpyxl
import re

#Read your Excel file
file_path = "results.xlsx" # Replace with your own file path.
workbook = openpyxl.load_workbook(file_path)
sheet = workbook.active

# Define the patterns to extract the desired text
start_pattern = "relationship_names%5D="
end_pattern = "&work_search%5Bfreeform_names"

# Iterate over the cells in Column A and modify their values
for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=1):
	cell = row[0]
	text = cell.value
	
	#Find the start and end indices of the desired text
	start_index = text.find(start_pattern)
	end_index = text.find(end_pattern)
	
	if start_index != -1 and end_index != -1:
		#Extract the desired text and update the cell values
		extracted_text = text[start_index + len(start_pattern):end_index]
		
		# Replace remaining symbols within pairings with the correct formatting
		extracted_text = extracted_text.replace('+', ' ').replace('%2F', '/').replace('%7C', '|').replace('%26', '&')

		cell.value = extracted_text
		
# Iterate over the celfls in Column B and remove " Found?"
for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=2, max_col=2):
	cell = row[0]
	text = cell.value
	
	if text is not None:	
		# Remove " Found?" from the cell
		text = text.replace("Found?", "")
		cell.value = text
		
# Save the modified Excel file
workbook.save("final_output.xlsx")
print("Process complete")
